# =============================================================================#
#                                                                              #
#           The uniques adaptive swarm particle optimization module            #
#           for the short-ranged ordering correction                           #
#                                                                              #
# -----------------------------------------------------------------------------#
# This module performs UAPSO using an external database.                      #
# -----------------------------------------------------------------------------#
# Original version: March 2022 by Okan K. Orhan                                #
# =============================================================================#


# !/bin/python3

# Importing the libraries
import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook

from Modulus.bravais_lattice_info import bravais
from Modulus.units_info import convert
from Modulus.output_info import out_write


# Function required to append to XLSX file for pandas version < 1.4
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return

    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    writer.book = load_workbook(filename)

    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    if truncate_sheet and sheet_name in writer.book.sheetnames:
        idx = writer.book.sheetnames.index(sheet_name)
        writer.book.remove(writer.book.worksheets[idx])
        writer.book.create_sheet(sheet_name, idx)

    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    writer.save()


# Function to calculate the objectives during UAPSO
def objectives(fileout, beta_list, inp_par, ext_data):
    obj = []
    obj_list = inp_par[10]
    obj_weight = inp_par[12]
    beta_limits = constraints(inp_par, ext_data)
    sub_num_sys = ext_data.shape[0] - 1

    if 1 in obj_list:
        # Gibbs free energy terms
        G0 = ext_data['Energy'].iloc[0]
        DeltaGmj = beta_list * (ext_data['Energy'].iloc[1:sub_num_sys + 1] - G0)
        # G_norm_factor = np.min(ext_data['Energy'])
        G_norm_factor = sum(beta_limits * (ext_data['Energy'].iloc[1:sub_num_sys + 1] - G0))

        # Elastic energy terms
        if "Bulk modulus" in ext_data.columns:
            tV0 = bravais(fileout, ext_data['Bravais'].iloc[0],
                          np.array(ext_data.filter(regex=r'^Celldm', axis=1).iloc[0])).volume
            tB0 = ext_data['Bulk modulus'].iloc[0]

            tibrav = np.array(ext_data['Bravais'].iloc[1:sub_num_sys + 1])
            tcelldm = np.array(ext_data.filter(regex=r'^Celldm', axis=1).iloc[1:sub_num_sys + 1])
            tV = []
            for i in range(sub_num_sys):
                tV.append(bravais(fileout, tibrav[i], tcelldm[i]).volume)
            tB = ext_data['Bulk modulus'].iloc[1:sub_num_sys + 1]
            tV = np.array(tV)

            DeltaUmj = beta_list * (np.abs(tV - tV0) * tB)
            DeltaUmj = pd.Series(DeltaUmj, index=DeltaGmj.index)

            tconv = 1.0
            if 'Bohr' in inp_par[5]:
                tconv *= convert().a02m_3
            elif 'Angstrom' in inp_par[5]:
                tconv *= convert().A2m

            if 'GPa' in inp_par[5]:
                tconv *= 1.0e9
            elif 'bar' in inp_par[5]:
                tconv *= convert().bar2Pa

            if 'eV' in inp_par[5]:
                tconv *= convert().J2eV
            elif 'Ha' in inp_par[5]:
                tconv *= convert().Ha2J
            elif 'Ry' in inp_par[5]:
                tconv *= convert().Ry2J

            DeltaGmj = DeltaGmj + tconv * DeltaUmj
            G_norm_factor += sum(tconv * beta_list * (np.abs(tV - tV0) * tB))

        DeltaG = G_norm_factor / np.sum(DeltaGmj+1.0e-6)
        if DeltaG < 0:
            DeltaG = 1.0e5

        obj.append(DeltaG)

    if 2 in obj_list:
        DeltaSize = 0.0
        if not ext_data['Bravais'].isin([4]).any():
            tlat = (1.0 - (ext_data['Celldm 1'].iloc[1:sub_num_sys + 1] / ext_data['Celldm 1'].iloc[0]))
            DeltaSize = np.sqrt(np.sum(beta_list * tlat * tlat))
        else:
            tV0 = bravais(fileout, ext_data['Bravais'].iloc[0],
                          np.array(ext_data.filter(regex=r'^Celldm', axis=1).iloc[0])).volume
            for i in range(sub_num_sys):
                tcelldm = np.array(ext_data.filter(regex=r'^Celldm', axis=1).iloc[i + 1])
                tibrav = ext_data['Bravais'].iloc[i + 1]
                tV = bravais(fileout, tibrav, tcelldm).volume
                DeltaSize += beta_list[i] * (1.0 - (tV / tV0)) * (1.0 - (tV / tV0))
            DeltaSize = np.sqrt(DeltaSize)
        obj.append(DeltaSize)

    if 3 in obj_list:
        tval = (1.0 - (ext_data['VEC'].iloc[1:sub_num_sys + 1] / ext_data['VEC'].iloc[0]))
        DeltaVEC = np.sqrt(np.sum(beta_list * tval * tval))
        obj.append(DeltaVEC)

    if 4 in obj_list:
        tval = (1.0 - (ext_data['Electronegativity'].iloc[1:sub_num_sys + 1] / ext_data['Electronegativity'].iloc[0]))
        DeltaChi = np.sqrt(np.sum(beta_list * tval * tval))
        obj.append(DeltaChi)

    return obj


# Function to determine the limits for individual beta_jm
def constraints(inp_par, ext_data):
    cons_ind = inp_par[13]
    sub_num_sys = ext_data.shape[0] - 1
    beta_limits = np.ones(sub_num_sys)

    c1 = []
    if 1 in cons_ind:
        tM = len(inp_par[2])
        for i in range(sub_num_sys):
            tm = np.count_nonzero(ext_data.iloc[i + 1][0:tM], axis=None)
            if beta_limits[i] > (1.0 / (tm * tM)):
                c1.append((1.0 / (tm * tM)) / beta_limits[i])
            else:
                c1.append(1)
        beta_limits = c1 * beta_limits

    c2 = 0.0
    if 2 in cons_ind:
        for i in range(len(inp_par[2])):
            pc = inp_par[2][i]
            tc = sum(beta_limits * ext_data[pc][1:sub_num_sys + 1])
            if tc > inp_par[3][i] and tc > c2:
                c2 = inp_par[3][1] / tc
        beta_limits = c2 * beta_limits

    c3 = 1.0
    if 3 in cons_ind:
        c3 = 1.0 / sum(beta_limits)
        beta_limits = c3 * beta_limits

    return beta_limits


# Function to calculate the cost for given objectives
def cost_func(fileout, obj, inp_par):
    obj_weight_fixed = inp_par[11]
    obj_weight = inp_par[12]

    cost = 0.0
    if obj_weight_fixed:
        for i in range(len(obj)):
            cost += obj_weight[i] * obj[i]
    else:
        out_write.error(fileout, 'On-fly objective weights are not implemented!')
    return cost


# Particle object
class particle:

    def __init__(self, fileout, inp_par, ext_data):
        num_sub_sys = ext_data.shape[0] - 1
        beta_limits = constraints(inp_par, ext_data)

        self.position = np.random.random(num_sub_sys)
        # Enforcing position limits
        self.position = np.maximum(self.position, np.zeros(num_sub_sys))
        self.position = np.minimum(self.position, beta_limits)

        self.velocity = np.random.random(num_sub_sys)
        self.obj = objectives(fileout, self.position, inp_par, ext_data)
        self.violation = np.maximum(np.zeros(num_sub_sys),
                                    np.sign(self.position - beta_limits))
        self.penalty = np.sum(self.violation) * 1.0
        self.cost = cost_func(fileout, self.obj, inp_par) + self.penalty
        self.feasible_sol = self.cost
        self.evolutionary_factor = np.random.random(1)  # It is evolutionary feedback parameter
        self.cognitive = np.random.random(1)
        self.social = np.random.random(1)


# Optimized solution object
class opt_sol:

    def __init__(self):
        self.beta = np.array([])
        self.cost = 1.0e10


# Function to calculate the optimized figure of merits with SRO correction
def fig_of_merit(fileout, beta_list, inp_par, ext_data):
    fom = []
    index = []
    obj_list = inp_par[10]
    sub_num_sys = ext_data.shape[0] - 1

    if 1 in obj_list:

        # Gibbs free energy terms
        G0 = ext_data['Energy'].iloc[0]
        DeltaGmj = beta_list * (ext_data['Energy'].iloc[1:sub_num_sys + 1] - G0)

        # Elastic energy terms
        tV0 = bravais(fileout, ext_data['Bravais'].iloc[0],
                      np.array(ext_data.filter(regex=r'^Celldm', axis=1).iloc[0])).volume
        tB0 = ext_data['Bulk modulus'].iloc[0]

        if "Bulk modulus" in ext_data.columns:
            DeltaUmj = []
            for i in range(sub_num_sys):
                tibrav = ext_data['Bravais'].iloc[i + 1]
                tcelldm = np.array(ext_data.filter(regex=r'^Celldm', axis=1).iloc[i + 1])
                tV = bravais(fileout, tibrav, tcelldm).volume
                tB = ext_data['Bulk modulus'].iloc[i + 1]
                DeltaUmj.append(beta_list[i] * (np.abs(tV - tV0) * tB))
            DeltaUmj = pd.Series(DeltaUmj, index=DeltaGmj.index)

            tconv = 1.0
            if 'Bohr' in inp_par[5]:
                tconv *= convert().a02m_3
            elif 'Angstrom' in inp_par[5]:
                tconv *= convert().A2m

            if 'GPa' in inp_par[5]:
                tconv *= 1.0e9
            elif 'bar' in inp_par[5]:
                tconv *= convert().bar2Pa

            if 'eV' in inp_par[5]:
                tconv *= convert().J2eV
            elif 'Ha' in inp_par[5]:
                tconv *= convert().Ha2J
            elif 'Ry' in inp_par[5]:
                tconv *= convert().Ry2J

            DeltaGmj = DeltaGmj + tconv * DeltaUmj

        G_SRO = np.sum(DeltaGmj)
        fom.append(G_SRO)
        index.append('G_SRO')

    if 2 in obj_list:
        DeltaSize = 0.0
        if not ext_data['Bravais'].isin([4]).any():
            tlat = (1.0 - (ext_data['Celldm 1'].iloc[1:sub_num_sys + 1] / ext_data['Celldm 1'].iloc[0]))
            DeltaSize = np.sqrt(np.sum(beta_list * tlat * tlat))
            index.append('Delta a')
        else:
            tV0 = bravais(fileout, ext_data['Bravais'].iloc[0],
                          np.array(ext_data.filter(regex=r'^Celldm', axis=1).iloc[0])).volume
            for i in range(sub_num_sys):
                tcelldm = np.array(ext_data.filter(regex=r'^Celldm', axis=1).iloc[i + 1])
                tibrav = ext_data['Bravais'].iloc[i+1]
                tV = bravais(fileout, tibrav, tcelldm).volume
                DeltaSize += beta_list[i] * (1.0 - (tV / tV0)) * (1.0 - (tV / tV0))
            DeltaSize = np.sqrt(DeltaSize)
            index.append('Delta V')
        fom.append(DeltaSize)

    if 3 in obj_list:
        tval = (1.0 - (ext_data['VEC'].iloc[1:sub_num_sys + 1] / ext_data['VEC'].iloc[0]))
        DeltaVEC = np.sqrt(np.sum(beta_list * tval * tval))
        fom.append(DeltaVEC)
        index.append('Delta VEC')

    if 4 in obj_list:
        tval = (1.0 - (ext_data['Electronegativity'].iloc[1:sub_num_sys + 1] / ext_data['Electronegativity'].iloc[0]))
        DeltaChi = np.sqrt(np.sum(beta_list * tval * tval))
        fom.append(DeltaChi)
        index.append('Delta Chi')

    return fom, index


# Main function for UAPSO
def UAPSO(fileout, inp_par, ext_data):
    # Simulations parameters from the input parameters
    max_run = inp_par[6]
    max_it = inp_par[7]
    pop_size = inp_par[8]
    num_sub_sys = ext_data.shape[0] - 1
    beta_limits = constraints(inp_par, ext_data)
    c_min = 0.0
    c_max = 4.0

    # Writing initial information into the output XLSX file
    file_name = inp_par[1] + '.xlsx'
    header = ext_data[['Name'] + inp_par[2] + ['Energy']].iloc[0:1]
    header.rename(columns={'Name': 'System'}, inplace=True)
    header.to_excel(file_name, sheet_name='Global best', index=None, float_format="%.4f", startrow=0, startcol=0)

    # Starting optimization runs
    global_best_run = opt_sol()  # Absolute global best after all runs

    for irun in range(max_run):

        global_best = opt_sol()  # Global best after iterations of the current run

        # Recording the global best through iterations for the current run if OptHistory = T
        if inp_par[9]:
            global_best_history_beta = pd.DataFrame()

        # Initialising the swarm for the current run
        out_write.iter(fileout, 2, '\n Run ', irun + 1, ' ... ')
        swarm = []
        swarm_best = []
        for ipop in range(pop_size):
            swarm.append(particle(fileout, inp_par, ext_data))
            swarm_best.append(particle(fileout, inp_par, ext_data))
            if swarm_best[ipop].cost < global_best.cost:
                global_best.beta = swarm_best[ipop].position
                global_best.cost = swarm_best[ipop].cost
        out_write.misc(fileout, 8, 'Initialization ')

        # Starting  optimization iteration for the current run
        max_it_reset = 0
        for iit in range(max_it):
            global_best_cost_pre = global_best.cost
            # Iterating over population
            for ipop in range(pop_size):
                # Determining the evolutionary factor
                if global_best.cost > 0:
                    if swarm[ipop].penalty == 0.0:
                        swarm[ipop].feasible_sol = swarm[ipop].cost
                    swarm[ipop].evolutionary_factor = (swarm_best[ipop].cost - global_best.cost) / swarm[
                        ipop].feasible_sol
                else:
                    swarm[ipop].evolutionary_factor = swarm_best[ipop].cost / global_best.cost

                # Determining the cognitive and social constants
                if swarm[ipop].evolutionary_factor <= 0.5:
                    swarm[ipop].cognitive = ((c_max - c_min) * (max_it - iit) / max_it) + c_min
                    swarm[ipop].social = c_max - ((c_max - c_min) * (max_it - iit) / max_it)
                elif swarm[ipop].evolutionary_factor > 0.5 and swarm[ipop].evolutionary_factor <= 1.0:
                    swarm[ipop].cognitive = c_max - ((c_max - c_min) * (max_it - iit) / max_it)
                    swarm[ipop].social = ((c_max - c_min) * (max_it - iit) / max_it) + c_min
                else:
                    swarm[ipop].cognitive = 0.5 * (c_max + c_min)
                    swarm[ipop].social = 0.5 * (c_max + c_min)

                # Updating particle velocities
                    # Re-initializing zero velocities
                swarm[ipop].velocity = swarm[ipop].velocity +\
                                          (1.0 - np.sign(swarm[ipop].velocity)) \
                                        * (1.0 + np.sign(swarm[ipop].velocity)) \
                                        * np.random.random(num_sub_sys) * beta_limits
                swarm[ipop].velocity = swarm[ipop].evolutionary_factor * swarm[ipop].velocity \
                                       + swarm[ipop].cognitive * np.random.random(num_sub_sys) \
                                       * (swarm_best[ipop].position - swarm[ipop].position) \
                                       + swarm[ipop].social * np.random.random(num_sub_sys) \
                                       * (global_best.beta - swarm[ipop].position) \
                                       - (1.0 - swarm[ipop].evolutionary_factor) \
                                       * (global_best.beta - swarm_best[ipop].position)
                # Enforcing velocity limits
                #swarm[ipop].velocity = np.maximum(swarm[ipop].velocity, np.zeros(num_sub_sys))
                #swarm[ipop].velocity = np.minimum(swarm[ipop].velocity, beta_limits)

                # Updating position
                swarm[ipop].position = 0.3 * swarm[ipop].position + 0.7 * swarm[ipop].velocity
                # Enforcing position limits
                swarm[ipop].position = np.maximum(swarm[ipop].position, np.zeros(num_sub_sys))
                swarm[ipop].position = np.minimum(swarm[ipop].position, beta_limits)

                # Evaluating cost and penalty
                swarm[ipop].obj = objectives(fileout, swarm[ipop].position, inp_par, ext_data)
                swarm[ipop].violation = np.maximum(np.zeros(num_sub_sys),
                                                   np.sign(beta_limits - swarm[ipop].position))
                swarm[ipop].penalty = np.sum(swarm[ipop].violation) * 1.0e5
                swarm[ipop].cost = cost_func(fileout, swarm[ipop].obj, inp_par)

                # Updating swarm best
                if swarm[ipop].cost < swarm_best[ipop].cost:
                    swarm_best[ipop] = swarm[ipop]

                # Updating global best for the current run
                if swarm_best[ipop].cost < global_best.cost:
                    global_best.beta = swarm_best[ipop].position
                    global_best.cost = swarm_best[ipop].cost
                    max_it_reset = 0
                    # Recording global best instances for the current run
                    if inp_par[9]:
                        tfom, tindex = fig_of_merit(fileout, global_best.beta, inp_par, ext_data)
                        tser = pd.Series(global_best.beta, index=ext_data['Name'].iloc[1:num_sub_sys + 1])
                        tser = tser.append(pd.Series(tfom, index=tindex))
                        global_best_history_beta['Ite ' + str(iit + 1)] = tser

            if global_best.cost == global_best_cost_pre:
                max_it_reset += 1
            out_write.iter(fileout, 10, 'Iteration ', iit + 1, ' --> Global Lowest Cost : ' + str(global_best.cost))

            # Breaking the iteration, if the global best of the current run does not change for 200 iteration
            if max_it_reset != 0 and max_it_reset % 200 == 0:
                out_write.misc(fileout, 8, 'WARNING: Global best was not updated in the last 200 iterations! '
                                           'UAPSO cycle for this run is terminated!')
                #break

        # Writing the global best history for the current run into the XLSX file
        if inp_par[9]:
            append_df_to_excel(file_name, global_best_history_beta, sheet_name='Beta in run ' + str(irun + 1),
                               startrow=0, float_format="%.4f")

        # Updating the absolute global best
        if global_best.cost < global_best_run.cost:
            global_best_run = global_best

    # Writing the absolute global best into the XLSX file
    out_write.misc(fileout, 0, '\n')
    out_write.misc(fileout, 0, 'Writing the final solution in ' + inp_par[1] + '.xlsx ...')

    final_data = ext_data[['Name'] + inp_par[2] + ['Energy']].iloc[1:num_sub_sys + 1]
    final_data.rename(columns={'Name': 'Sub-system'}, inplace=True)
    final_data['Beta limits'] = pd.Series(beta_limits, index=ext_data.index[1:num_sub_sys + 1])
    final_data['Beta_jm'] = pd.Series(global_best_run.beta, index=ext_data.index[1:num_sub_sys + 1])
    append_df_to_excel(file_name, final_data, sheet_name='Global best',
                       startrow=3, index=None, float_format="%.4f")
    fom_final, fom_index = fig_of_merit(fileout, global_best_run.beta, inp_par, ext_data)
    fom_final = pd.Series(fom_final, name='FoM', index=fom_index)
    append_df_to_excel(file_name, fom_final, sheet_name='Global best',
                       startrow=num_sub_sys + 5, startcol=0, float_format="%.4f")
